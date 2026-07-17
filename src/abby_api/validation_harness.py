from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from statistics import mean
from uuid import uuid4

from openpyxl import load_workbook

from abby_api.cdr_calibration_runner import generate_cdr_calibration_artifacts
from abby_api.repositories.memory import get_prediction, new_project, save_structure
from abby_api.schemas.predictions import PredictionRequest
from abby_api.schemas.structures import (
    ChainMapping,
    StructureInput,
    StructureValidationRequest,
)
from abby_api.services.cdr_stress_harness import run_cdr_mutation_stress_batch
from abby_api.services.predictions import create_prediction
from abby_api.services.simulation import (
    SimulationRunConfig,
    run_gromacs_cif_simulation,
)
from abby_api.services.structure_parsing import (
    convert_pdb_to_mmcif,
    parse_structure_file,
    summarize_structure,
)
from abby_api.services.structures import validate_structure

DEFAULT_DATASET_ROOT = Path("validation_dataset/ANDD_pdb")
DEFAULT_WORKBOOK_NAME = "Antibody and Nanobody Design Dataset (ANDD)_v2.xlsx"
DEFAULT_OUTPUT_ROOT = Path("data/validation_runs/andd")

_MISSING_VALUES = {"", "na", "n/a", "none", "null", "\\", "-"}
_CHAIN_SEPARATORS = re.compile(r"[;,/|]+|\s+")
_FLOAT_RE = re.compile(r"[-+]?(?:\d+\.\d*|\.\d+|\d+)(?:[eE][-+]?\d+)?")
_CDR_QUALITY_BLOCKING_WARNING_CODES = {
    "CDR_CHAIN_ROLE_AMBIGUOUS",
    "CDR_BOUNDARY_AMBIGUOUS",
    "CDR_MOTIF_FALLBACK_USED",
    "CDR_NUMBERING_MISSING",
}


@dataclass
class AnddReference:
    row_index: int
    source: str
    pdb_id: str
    affinity_kd_m: float | None
    delta_g_kj_mol: float | None
    affinity_method: str | None
    reason_code: str | None
    predicted_or_not: str | None
    provenance: str | None
    heavy_chain_auth_ids: list[str] = field(default_factory=list)
    light_chain_auth_ids: list[str] = field(default_factory=list)
    antigen_chain_auth_ids: list[str] = field(default_factory=list)

    def primary_label_kind(self) -> str | None:
        if self.affinity_kd_m is not None:
            return "kd_m"
        if self.delta_g_kj_mol is not None:
            return "delta_g_kj_mol"
        return None

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


@dataclass
class ValidationCaseResult:
    pdb_id: str
    source_pdb_path: str
    converted_mmcif_path: str | None = None
    structure_id: str | None = None
    prediction_id: str | None = None
    status: str = "pending"
    conversion_status: str = "pending"
    validation_status: str = "pending"
    prediction_status: str = "pending"
    simulation_status: str = "skipped"
    md_handoff_ready: bool = False
    gromacs_cif_available: bool = False
    ground_truth_row_count: int = 0
    reference_source: str | None = None
    reference_row_index: int | None = None
    reference_kind: str | None = None
    experimental_kd_m: float | None = None
    experimental_log_k: float | None = None
    experimental_delta_g_kj_mol: float | None = None
    experimental_delta_g_kcal_mol: float | None = None
    predicted_log_k: float | None = None
    predicted_delta_g_kcal_mol: float | None = None
    cdr_quality_baseline_score: float | None = None
    cdr_quality_baseline_class: str | None = None
    cdr_quality_baseline_drift_flag: bool | None = None
    cdr_quality_baseline_observed_pass: bool | None = None
    validation_warnings: list[str] = field(default_factory=list)
    validation_errors: list[str] = field(default_factory=list)
    prediction_warnings: list[str] = field(default_factory=list)
    simulation_notes: list[str] = field(default_factory=list)
    error: str | None = None

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


@dataclass(frozen=True)
class ValidationMetrics:
    paired_cases: int
    pearson_log_k: float | None
    spearman_log_k: float | None
    mae_log_k: float | None
    rmse_log_k: float | None
    pearson_delta_g_kcal_mol: float | None
    spearman_delta_g_kcal_mol: float | None
    mae_delta_g_kcal_mol: float | None
    rmse_delta_g_kcal_mol: float | None

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


@dataclass(frozen=True)
class ValidationReport:
    generated_at: str
    dataset_root: str
    workbook_path: str
    output_dir: str
    project_id: str
    total_structures: int
    converted_structures: int
    validated_structures: int
    predicted_structures: int
    simulation_attempted_structures: int
    simulation_available_structures: int
    simulation_completed_structures: int
    matched_structures: int
    failed_structures: int
    metrics: ValidationMetrics
    cases: list[ValidationCaseResult]

    def to_dict(self) -> dict[str, object]:
        return {
            "generated_at": self.generated_at,
            "dataset_root": self.dataset_root,
            "workbook_path": self.workbook_path,
            "output_dir": self.output_dir,
            "project_id": self.project_id,
            "total_structures": self.total_structures,
            "converted_structures": self.converted_structures,
            "validated_structures": self.validated_structures,
            "predicted_structures": self.predicted_structures,
            "simulation_attempted_structures": self.simulation_attempted_structures,
            "simulation_available_structures": self.simulation_available_structures,
            "simulation_completed_structures": self.simulation_completed_structures,
            "matched_structures": self.matched_structures,
            "failed_structures": self.failed_structures,
            "metrics": self.metrics.to_dict(),
            "cases": [case.to_dict() for case in self.cases],
        }


def _normalize_text(value: object | None) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"[^a-z0-9]+", "_", text.strip().lower()).strip("_")


def _text_or_none(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_float(value: object | None) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text or text.lower() in _MISSING_VALUES:
        return None
    match = _FLOAT_RE.search(text.replace(",", ""))
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _parse_chain_ids(value: object | None) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text or text.lower() in _MISSING_VALUES:
        return []
    chains: list[str] = []
    for part in _CHAIN_SEPARATORS.split(text):
        token = part.strip()
        if not token or token.lower() in _MISSING_VALUES:
            continue
        if token not in chains:
            chains.append(token)
    return chains


def _header_lookup(headers: list[object]) -> dict[str, int]:
    return {_normalize_text(header): index for index, header in enumerate(headers)}


def _find_column(lookup: dict[str, int], *aliases: str) -> int | None:
    normalized_lookup = list(lookup.items())
    for alias in aliases:
        alias_norm = _normalize_text(alias)
        for header_norm, index in normalized_lookup:
            if (
                header_norm == alias_norm
                or header_norm.startswith(alias_norm)
                or alias_norm in header_norm
            ):
                return index
    return None


def _cell(row: tuple[object, ...], index: int | None) -> object | None:
    if index is None or index >= len(row):
        return None
    return row[index]


def _load_reference_rows(workbook_path: Path) -> list[AnddReference]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    lookup = _header_lookup(headers)

    source_idx = _find_column(lookup, "Source")
    pdb_id_idx = _find_column(lookup, "PDB_ID")
    heavy_idx = _find_column(lookup, "H_Chain Auth Asym ID")
    light_idx = _find_column(lookup, "L_Chain Auth Asym ID")
    antigen_idx = _find_column(lookup, "Ag_Auth Asym ID")
    kd_idx = _find_column(lookup, "Affinity_Kd(M)", "Affinity_Kd (M)")
    delta_g_idx = _find_column(lookup, "ΔGbinding", "∆Gbinding", "Gbinding")
    method_idx = _find_column(lookup, "Affinity_Method")
    reason_idx = _find_column(lookup, "Reason_Code")
    predicted_idx = _find_column(lookup, "Predicted_or_Not")
    provenance_idx = _find_column(lookup, "Provenance")

    references: list[AnddReference] = []
    for row_index, row in enumerate(rows, start=2):
        pdb_id = str(_cell(row, pdb_id_idx) or "").strip().upper()
        if not pdb_id:
            continue
        references.append(
            AnddReference(
                row_index=row_index,
                source=_text_or_none(_cell(row, source_idx)) or "",
                pdb_id=pdb_id,
                affinity_kd_m=_parse_float(_cell(row, kd_idx)),
                delta_g_kj_mol=_parse_float(_cell(row, delta_g_idx)),
                affinity_method=_text_or_none(_cell(row, method_idx)),
                reason_code=_text_or_none(_cell(row, reason_idx)),
                predicted_or_not=_text_or_none(_cell(row, predicted_idx)),
                provenance=_text_or_none(_cell(row, provenance_idx)),
                heavy_chain_auth_ids=_parse_chain_ids(_cell(row, heavy_idx)),
                light_chain_auth_ids=_parse_chain_ids(_cell(row, light_idx)),
                antigen_chain_auth_ids=_parse_chain_ids(_cell(row, antigen_idx)),
            )
        )
    workbook.close()
    return references


def _group_references_by_pdb_id(references: list[AnddReference]) -> dict[str, list[AnddReference]]:
    grouped: dict[str, list[AnddReference]] = {}
    for reference in references:
        grouped.setdefault(reference.pdb_id, []).append(reference)
    return grouped


def _reference_is_primary(reference: AnddReference) -> bool:
    if reference.predicted_or_not and "predicted" in reference.predicted_or_not.lower():
        return False
    return reference.affinity_kd_m is not None or reference.delta_g_kj_mol is not None


def _select_primary_reference(references: list[AnddReference]) -> AnddReference | None:
    if not references:
        return None
    experimental = [reference for reference in references if _reference_is_primary(reference)]
    if experimental:
        return experimental[0]
    numeric = [reference for reference in references if reference.primary_label_kind() is not None]
    if numeric:
        return numeric[0]
    return references[0]


def _chain_mapping_from_reference(
    reference: AnddReference | None, available_chains: list[str]
) -> ChainMapping:
    available = [chain.strip() for chain in available_chains if chain.strip()]
    if reference is None:
        partner_1 = available[:1]
        partner_2 = available[1:2] if len(available) > 1 else []
        return ChainMapping(partner_1=partner_1, partner_2=partner_2)

    partner_1 = [*reference.heavy_chain_auth_ids, *reference.light_chain_auth_ids]
    partner_2 = list(reference.antigen_chain_auth_ids)

    partner_1 = [chain for chain in partner_1 if chain in available]
    partner_2 = [chain for chain in partner_2 if chain in available and chain not in partner_1]

    if not partner_1 and available:
        partner_1 = available[:1]
    if not partner_2:
        remaining = [chain for chain in available if chain not in partner_1]
        partner_2 = remaining[:1]
    if not partner_2 and len(available) > 1:
        partner_2 = available[1:2]

    return ChainMapping(partner_1=partner_1, partner_2=partner_2)


def _experimental_log_k(
    reference: AnddReference, temperature_kelvin: float = 298.15
) -> float | None:
    if reference.affinity_kd_m is not None and reference.affinity_kd_m > 0:
        return -math.log10(reference.affinity_kd_m)
    if reference.delta_g_kj_mol is not None:
        r_kj_per_mol_k = 0.008314462618
        return -reference.delta_g_kj_mol / (2.303 * r_kj_per_mol_k * temperature_kelvin)
    return None


def _experimental_delta_g_kcal_mol(
    reference: AnddReference, temperature_kelvin: float = 298.15
) -> float | None:
    log_k = _experimental_log_k(reference, temperature_kelvin=temperature_kelvin)
    if log_k is None:
        return None
    r_kcal_per_mol_k = 0.00198720425864083
    return -(2.303 * r_kcal_per_mol_k * temperature_kelvin * log_k)


def _pearson(values_x: list[float], values_y: list[float]) -> float | None:
    if len(values_x) < 2 or len(values_y) < 2:
        return None
    mean_x = mean(values_x)
    mean_y = mean(values_y)
    numerator = sum((x - mean_x) * (y - mean_y) for x, y in zip(values_x, values_y))
    denominator_x = math.sqrt(sum((x - mean_x) ** 2 for x in values_x))
    denominator_y = math.sqrt(sum((y - mean_y) ** 2 for y in values_y))
    denominator = denominator_x * denominator_y
    if denominator == 0:
        return None
    return numerator / denominator


def _average_ranks(values: list[float]) -> list[float]:
    indexed = sorted(enumerate(values), key=lambda item: item[1])
    ranks = [0.0] * len(values)
    position = 0
    while position < len(indexed):
        end = position
        while end + 1 < len(indexed) and indexed[end + 1][1] == indexed[position][1]:
            end += 1
        rank = (position + 1 + end + 1) / 2
        for slot in range(position, end + 1):
            ranks[indexed[slot][0]] = rank
        position = end + 1
    return ranks


def _spearman(values_x: list[float], values_y: list[float]) -> float | None:
    if len(values_x) < 2 or len(values_y) < 2:
        return None
    return _pearson(_average_ranks(values_x), _average_ranks(values_y))


def _mae(values_x: list[float], values_y: list[float]) -> float | None:
    if not values_x:
        return None
    return sum(abs(x - y) for x, y in zip(values_x, values_y)) / len(values_x)


def _rmse(values_x: list[float], values_y: list[float]) -> float | None:
    if not values_x:
        return None
    return math.sqrt(sum((x - y) ** 2 for x, y in zip(values_x, values_y)) / len(values_x))


def _compute_metrics(cases: list[ValidationCaseResult]) -> ValidationMetrics:
    log_pairs = [
        (case.predicted_log_k, case.experimental_log_k)
        for case in cases
        if case.predicted_log_k is not None and case.experimental_log_k is not None
    ]
    delta_pairs = [
        (case.predicted_delta_g_kcal_mol, case.experimental_delta_g_kcal_mol)
        for case in cases
        if (
            case.predicted_delta_g_kcal_mol is not None
            and case.experimental_delta_g_kcal_mol is not None
        )
    ]

    log_x = [item[0] for item in log_pairs]
    log_y = [item[1] for item in log_pairs]
    delta_x = [item[0] for item in delta_pairs]
    delta_y = [item[1] for item in delta_pairs]

    return ValidationMetrics(
        paired_cases=len(log_pairs),
        pearson_log_k=_pearson(log_x, log_y),
        spearman_log_k=_spearman(log_x, log_y),
        mae_log_k=_mae(log_x, log_y),
        rmse_log_k=_rmse(log_x, log_y),
        pearson_delta_g_kcal_mol=_pearson(delta_x, delta_y),
        spearman_delta_g_kcal_mol=_spearman(delta_x, delta_y),
        mae_delta_g_kcal_mol=_mae(delta_x, delta_y),
        rmse_delta_g_kcal_mol=_rmse(delta_x, delta_y),
    )


def _write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True, default=str), encoding="utf-8")


def _extract_cdr_quality_baseline(
    prediction: object,
) -> tuple[float | None, str | None, bool | None]:
    provenance = getattr(prediction, "provenance", None)
    cdr_annotation = getattr(provenance, "cdr_annotation", None)
    quality_baseline = getattr(cdr_annotation, "quality_baseline", None)
    if quality_baseline is None:
        return None, None, None

    score_value = getattr(quality_baseline, "score", None)
    score = float(score_value) if isinstance(score_value, (int, float)) else None
    classification = getattr(quality_baseline, "classification", None)
    drift_flag = getattr(quality_baseline, "drift_flag", None)
    return score, classification, drift_flag if isinstance(drift_flag, bool) else None


def _cdr_quality_observed_pass(validation: object) -> bool:
    errors = list(getattr(validation, "errors", []) or [])
    if errors:
        return False
    warnings = set(getattr(validation, "warnings", []) or [])
    return not any(code in warnings for code in _CDR_QUALITY_BLOCKING_WARNING_CODES)


def _serialize_case_row(case: ValidationCaseResult) -> dict[str, object]:
    return {
        "pdb_id": case.pdb_id,
        "source_pdb_path": case.source_pdb_path,
        "converted_mmcif_path": case.converted_mmcif_path or "",
        "structure_id": case.structure_id or "",
        "prediction_id": case.prediction_id or "",
        "status": case.status,
        "conversion_status": case.conversion_status,
        "validation_status": case.validation_status,
        "prediction_status": case.prediction_status,
        "simulation_status": case.simulation_status,
        "md_handoff_ready": str(case.md_handoff_ready),
        "gromacs_cif_available": str(case.gromacs_cif_available),
        "ground_truth_row_count": case.ground_truth_row_count,
        "reference_source": case.reference_source or "",
        "reference_row_index": case.reference_row_index or "",
        "reference_kind": case.reference_kind or "",
        "experimental_kd_m": "" if case.experimental_kd_m is None else case.experimental_kd_m,
        "experimental_log_k": "" if case.experimental_log_k is None else case.experimental_log_k,
        "experimental_delta_g_kj_mol": ""
        if case.experimental_delta_g_kj_mol is None
        else case.experimental_delta_g_kj_mol,
        "experimental_delta_g_kcal_mol": ""
        if case.experimental_delta_g_kcal_mol is None
        else case.experimental_delta_g_kcal_mol,
        "predicted_log_k": "" if case.predicted_log_k is None else case.predicted_log_k,
        "predicted_delta_g_kcal_mol": ""
        if case.predicted_delta_g_kcal_mol is None
        else case.predicted_delta_g_kcal_mol,
        "cdr_quality_baseline_score": ""
        if case.cdr_quality_baseline_score is None
        else case.cdr_quality_baseline_score,
        "cdr_quality_baseline_class": case.cdr_quality_baseline_class or "",
        "cdr_quality_baseline_drift_flag": ""
        if case.cdr_quality_baseline_drift_flag is None
        else str(case.cdr_quality_baseline_drift_flag),
        "cdr_quality_baseline_observed_pass": ""
        if case.cdr_quality_baseline_observed_pass is None
        else str(case.cdr_quality_baseline_observed_pass),
        "error": case.error or "",
    }


def _build_structure_input(converted_path: Path) -> tuple[StructureInput, str]:
    structure_id = uuid4()
    payload = converted_path.read_bytes()
    structure = StructureInput(
        structure_id=structure_id,
        format="mmcif",
        source="derived",
        filename=converted_path.name,
        sha256=hashlib.sha256(payload).hexdigest(),
        mode="antibody_antigen",
        chains=None,
    )
    return structure, structure_id.hex


def _build_cdr_stress_resilience_assertions(
    *,
    parsed_specs: int,
    failed_specs: int,
    total_specs: int,
    parsed_spec_chains: list[str],
    structure_available_chains: list[str],
) -> dict[str, dict[str, object]]:
    if total_specs <= 0:
        failure_rate = 0.0
    else:
        failure_rate = failed_specs / total_specs
    structure_chain_set = {chain.strip() for chain in structure_available_chains if chain.strip()}
    missing_spec_chains = sorted(
        {
            chain.strip()
            for chain in parsed_spec_chains
            if chain.strip() and chain.strip() not in structure_chain_set
        }
    )

    return {
        "nonzero_parse_success": {
            "passed": parsed_specs > 0,
            "observed": parsed_specs,
            "expected": ">=1",
        },
        "failure_rate_within_limit": {
            "passed": failure_rate <= (1 / 3),
            "observed": failure_rate,
            "expected": "<=0.3333333333",
        },
        "spec_chains_present_in_structures": {
            "passed": len(missing_spec_chains) == 0,
            "observed_missing_chains": missing_spec_chains,
            "expected": "all parsed spec chains present in processed structures",
        },
    }


def run_andd_validation_harness(
    *,
    dataset_root: Path = DEFAULT_DATASET_ROOT,
    workbook_path: Path | None = None,
    output_dir: Path = DEFAULT_OUTPUT_ROOT,
    pdb_ids: list[str] | None = None,
    limit: int | None = None,
    simulation_policy: str = "skip",
    temperature_kelvin: float = 298.15,
    cdr_stress_specs: list[str] | None = None,
) -> ValidationReport:
    workbook_path = workbook_path or dataset_root / DEFAULT_WORKBOOK_NAME
    output_dir = output_dir.resolve()
    converted_dir = output_dir / "converted_mmcif"
    report_dir = output_dir / "reports"
    report_dir.mkdir(parents=True, exist_ok=True)

    references = _load_reference_rows(workbook_path)
    grouped_references = _group_references_by_pdb_id(references)

    all_pdb_paths = sorted((dataset_root / "All_structures").glob("*.pdb"))
    if pdb_ids:
        wanted = {pdb_id.upper() for pdb_id in pdb_ids}
        all_pdb_paths = [path for path in all_pdb_paths if path.stem.upper() in wanted]
    if limit is not None:
        all_pdb_paths = all_pdb_paths[:limit]

    project = new_project(f"ANDD validation {datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}")
    cases: list[ValidationCaseResult] = []
    structure_chain_universe: set[str] = set()

    for pdb_path in all_pdb_paths:
        pdb_id = pdb_path.stem.upper()
        references_for_pdb = grouped_references.get(pdb_id, [])
        converted_path = converted_dir / f"{pdb_id}.mmcif"
        case = ValidationCaseResult(
            pdb_id=pdb_id,
            source_pdb_path=str(pdb_path),
            converted_mmcif_path=str(converted_path),
            ground_truth_row_count=len(references_for_pdb),
        )

        try:
            convert_pdb_to_mmcif(pdb_path, converted_path)
            case.conversion_status = "completed"

            parsed_structure, parser_name = parse_structure_file(converted_path, "mmcif")
            summary = summarize_structure(
                parsed_structure,
                parser_name,
                file_path=converted_path,
                format_name="mmcif",
                prediction_mode="antibody_antigen",
            )
            structure_chain_universe.update(
                chain.strip() for chain in summary.available_chains if chain.strip()
            )
            structure_input, _ = _build_structure_input(converted_path)
            save_structure(structure_input, file_path=converted_path, summary=summary)
            case.structure_id = str(structure_input.structure_id)

            reference = _select_primary_reference(references_for_pdb)
            if reference is not None:
                case.reference_source = reference.source
                case.reference_row_index = reference.row_index
                case.reference_kind = reference.primary_label_kind()
                case.experimental_kd_m = reference.affinity_kd_m
                case.experimental_log_k = _experimental_log_k(reference, temperature_kelvin)
                case.experimental_delta_g_kj_mol = reference.delta_g_kj_mol
                case.experimental_delta_g_kcal_mol = _experimental_delta_g_kcal_mol(
                    reference, temperature_kelvin
                )
            chain_mapping = _chain_mapping_from_reference(reference, summary.available_chains)
            validation = validate_structure(
                StructureValidationRequest(
                    structure_id=structure_input.structure_id,
                    mode="antibody_antigen",
                    chains=chain_mapping,
                )
            )
            case.validation_status = "completed" if validation.valid else "failed"
            case.md_handoff_ready = bool(validation.md_handoff.get("ready_for_md_handoff", False))
            case.validation_warnings = list(validation.warnings)
            case.validation_errors = list(validation.errors)

            if not validation.valid:
                error_message = ", ".join(validation.errors) or "unknown error"
                raise RuntimeError(f"Validation failed for {pdb_id}: {error_message}")

            queued_prediction = create_prediction(
                PredictionRequest(
                    project_id=project.project_id,
                    mode="antibody_antigen",
                    structure_id=structure_input.structure_id,
                )
            )
            case.prediction_id = str(queued_prediction.prediction_id)
            persisted_prediction = get_prediction(queued_prediction.prediction_id)
            if persisted_prediction is None or persisted_prediction.consensus is None:
                raise RuntimeError("Prediction result could not be retrieved from memory store.")

            case.predicted_log_k = persisted_prediction.consensus.log_k
            case.predicted_delta_g_kcal_mol = persisted_prediction.consensus.delta_g_kcal_mol
            case.prediction_status = persisted_prediction.status
            (
                case.cdr_quality_baseline_score,
                case.cdr_quality_baseline_class,
                case.cdr_quality_baseline_drift_flag,
            ) = _extract_cdr_quality_baseline(persisted_prediction)
            case.cdr_quality_baseline_observed_pass = _cdr_quality_observed_pass(validation)

            if simulation_policy != "skip":
                simulation_result = run_gromacs_cif_simulation(
                    converted_path,
                    SimulationRunConfig(structure_format="mmcif"),
                    prediction_id=queued_prediction.prediction_id,
                    project_id=project.project_id,
                )
                case.simulation_status = (
                    "completed" if simulation_result.gromacs_available else "stubbed"
                )
                case.gromacs_cif_available = simulation_result.gromacs_available
                case.simulation_notes = list(simulation_result.notes)
            else:
                case.simulation_status = "skipped"

            case.status = "completed"
        except Exception as exc:  # pragma: no cover - captured per-structure for auditability
            case.status = "failed"
            case.error = str(exc)

        cases.append(case)

    metrics = _compute_metrics(cases)
    report = ValidationReport(
        generated_at=datetime.now(timezone.utc).isoformat(),
        dataset_root=str(dataset_root.resolve()),
        workbook_path=str(workbook_path.resolve()),
        output_dir=str(output_dir),
        project_id=str(project.project_id),
        total_structures=len(all_pdb_paths),
        converted_structures=sum(1 for case in cases if case.conversion_status == "completed"),
        validated_structures=sum(1 for case in cases if case.validation_status == "completed"),
        predicted_structures=sum(1 for case in cases if case.prediction_status == "completed"),
        simulation_attempted_structures=sum(
            1 for case in cases if case.simulation_status != "skipped"
        ),
        simulation_available_structures=sum(1 for case in cases if case.gromacs_cif_available),
        simulation_completed_structures=sum(
            1 for case in cases if case.simulation_status == "completed"
        ),
        matched_structures=sum(
            1
            for case in cases
            if case.reference_row_index is not None and case.status == "completed"
        ),
        failed_structures=sum(1 for case in cases if case.status == "failed"),
        metrics=metrics,
        cases=cases,
    )

    report_path = report_dir / "validation_report.json"
    _write_json(report_path, report.to_dict())
    case_rows = [_serialize_case_row(case) for case in cases]
    _write_csv(
        report_dir / "validation_cases.csv",
        case_rows,
        fieldnames=list(case_rows[0].keys()) if case_rows else [],
    )
    _write_csv(
        report_dir / "validation_manifest.csv",
        [
            {
                "pdb_id": case.pdb_id,
                "source_pdb_path": case.source_pdb_path,
                "converted_mmcif_path": case.converted_mmcif_path or "",
                "structure_id": case.structure_id or "",
                "prediction_id": case.prediction_id or "",
                "ground_truth_row_count": case.ground_truth_row_count,
                "status": case.status,
            }
            for case in cases
        ],
        fieldnames=[
            "pdb_id",
            "source_pdb_path",
            "converted_mmcif_path",
            "structure_id",
            "prediction_id",
            "ground_truth_row_count",
            "status",
        ],
    )

    try:
        generate_cdr_calibration_artifacts(report_path)
    except Exception as exc:  # pragma: no cover - best-effort artifact augmentation
        print(f"Failed to generate CDR calibration artifacts: {exc}")

    if cdr_stress_specs:
        stress_summary = run_cdr_mutation_stress_batch(cdr_stress_specs)
        parsed_spec_chains = [
            result.parsed_spec.chain_id
            for result in stress_summary.results
            if result.parsed_spec is not None
        ]
        resilience_assertions = _build_cdr_stress_resilience_assertions(
            parsed_specs=stress_summary.parsed_specs,
            failed_specs=stress_summary.failed_specs,
            total_specs=stress_summary.total_specs,
            parsed_spec_chains=parsed_spec_chains,
            structure_available_chains=sorted(structure_chain_universe),
        )
        _write_json(
            report_dir / "cdr_mutation_stress_report.json",
            {
                "total_specs": stress_summary.total_specs,
                "parsed_specs": stress_summary.parsed_specs,
                "failed_specs": stress_summary.failed_specs,
                "resilience_assertions": resilience_assertions,
                "results": [
                    {
                        "input_spec": result.input_spec,
                        "status": result.status,
                        "parsed_spec": (
                            {
                                "chain_id": result.parsed_spec.chain_id,
                                "start_seq_id": result.parsed_spec.start_seq_id,
                                "end_seq_id": result.parsed_spec.end_seq_id,
                                "insertion_code": result.parsed_spec.insertion_code,
                                "from_residue": result.parsed_spec.from_residue,
                                "to_residue": result.parsed_spec.to_residue,
                                "mode": result.parsed_spec.mode,
                            }
                            if result.parsed_spec is not None
                            else None
                        ),
                        "error": result.error,
                    }
                    for result in stress_summary.results
                ],
            },
        )

    return report


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Run the ANDD validation harness.")
    parser.add_argument("--dataset-root", type=Path, default=DEFAULT_DATASET_ROOT)
    parser.add_argument("--workbook-path", type=Path, default=None)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument("--pdb-id", action="append", dest="pdb_ids")
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument(
        "--simulation-policy",
        choices=["skip", "run_if_available", "force"],
        default="skip",
    )
    parser.add_argument(
        "--cdr-stress-spec",
        action="append",
        dest="cdr_stress_specs",
        help=(
            "Optional CDR mutation stress specification (repeatable). "
            "When provided, emits reports/cdr_mutation_stress_report.json."
        ),
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    report = run_andd_validation_harness(
        dataset_root=args.dataset_root,
        workbook_path=args.workbook_path,
        output_dir=args.output_dir,
        pdb_ids=args.pdb_ids,
        limit=args.limit,
        simulation_policy=args.simulation_policy,
        cdr_stress_specs=args.cdr_stress_specs,
    )
    print(json.dumps(report.to_dict(), indent=2, sort_keys=True, default=str))
    return 0


if __name__ == "__main__":  # pragma: no cover - manual execution entry point
    raise SystemExit(main())