from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

from abby_api.validation_harness import (
    DEFAULT_DATASET_ROOT,
    DEFAULT_OUTPUT_ROOT,
    DEFAULT_WORKBOOK_NAME,
    run_andd_validation_harness,
)


def _smoke_pdb_ids(dataset_root: Path, workbook_path: Path, smoke_limit: int) -> list[str]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    header_lookup = {
        str(header).strip().lower().replace(" ", "_"): index
        for index, header in enumerate(headers)
    }

    pdb_idx = header_lookup.get("pdb_id")
    antigen_idx = header_lookup.get("ag_auth_asym_id")
    predicted_idx = header_lookup.get("predicted_or_not")

    candidates: list[str] = []
    for row in rows:
        if pdb_idx is None or pdb_idx >= len(row):
            continue
        pdb_id = str(row[pdb_idx] or "").strip().upper()
        if not pdb_id:
            continue
        antigen_value = ""
        if antigen_idx is not None and antigen_idx < len(row):
            antigen_value = str(row[antigen_idx] or "").strip().lower()
        predicted_value = ""
        if predicted_idx is not None and predicted_idx < len(row):
            predicted_value = str(row[predicted_idx] or "").strip().lower()

        if antigen_value and antigen_value not in {"n/a", "na", "\\"} and predicted_value != "predicted":
            candidates.append(pdb_id)

    workbook.close()

    if not candidates:
        paths = sorted((dataset_root / "All_structures").glob("*.pdb"))
        return [path.stem.upper() for path in paths[:smoke_limit]]

    unique_candidates: list[str] = []
    for pdb_id in candidates:
        if pdb_id not in unique_candidates:
            unique_candidates.append(pdb_id)
    return unique_candidates[:smoke_limit]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Run the Abby ANDD validation harness with smoke/full modes."
    )
    parser.add_argument("--dataset-root", type=Path, default=DEFAULT_DATASET_ROOT)
    parser.add_argument(
        "--workbook-path",
        type=Path,
        default=DEFAULT_DATASET_ROOT / DEFAULT_WORKBOOK_NAME,
    )
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument(
        "--simulation-policy",
        choices=["skip", "run_if_available", "force"],
        default="skip",
    )
    parser.add_argument(
        "--smoke",
        action="store_true",
        help="Run a filtered smoke pass using the first few ANDD structures.",
    )
    parser.add_argument(
        "--smoke-limit",
        type=int,
        default=12,
        help="Number of structures to include in smoke mode.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Cap the number of structures processed after filtering.",
    )
    parser.add_argument(
        "--pdb-id",
        action="append",
        dest="pdb_ids",
        help="Limit the run to one or more specific PDB IDs.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    pdb_ids = list(args.pdb_ids or [])
    if args.smoke and not pdb_ids:
        pdb_ids = _smoke_pdb_ids(args.dataset_root, args.workbook_path, args.smoke_limit)

    report = run_andd_validation_harness(
        dataset_root=args.dataset_root,
        workbook_path=args.workbook_path,
        output_dir=args.output_dir,
        pdb_ids=pdb_ids or None,
        limit=args.limit,
        simulation_policy=args.simulation_policy,
    )
    print(json.dumps(report.to_dict(), indent=2, sort_keys=True, default=str))
    return 0


if __name__ == "__main__":  # pragma: no cover - convenience CLI wrapper
    raise SystemExit(main())